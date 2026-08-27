"""Build the PDF (via reportlab) and Excel (via openpyxl) downloads for the
Visual Double Open Clusters observing list.

Output: observing-lists/downloads/visual-double-open-clusters.{pdf,xlsx}

Usage: python scripts/build_visual_doubles_downloads.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _visual_doubles_data import PAIRS  # noqa: E402

SITE_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = SITE_ROOT / "observing-lists" / "downloads"
OUT_DIR.mkdir(parents=True, exist_ok=True)

PDF_OUT = OUT_DIR / "visual-double-open-clusters.pdf"
XLSX_OUT = OUT_DIR / "visual-double-open-clusters.xlsx"


# ------------------------------------------------------------
# PDF: reportlab platypus, landscape A4, single wide table
# ------------------------------------------------------------
def build_pdf() -> None:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontSize=18, spaceAfter=8, alignment=0,
    )
    subtitle_style = ParagraphStyle(
        "subtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, spaceAfter=16,
    )
    intro_style = ParagraphStyle(
        "intro", parent=styles["Normal"],
        fontSize=10, leading=13, spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        "cell", parent=styles["Normal"],
        fontSize=8, leading=10,
    )
    cell_bold = ParagraphStyle(
        "cell_bold", parent=cell_style,
        fontName="Helvetica-Bold",
    )
    header_style = ParagraphStyle(
        "header", parent=styles["Normal"],
        fontSize=8.5, leading=10, fontName="Helvetica-Bold",
        textColor=colors.whitesmoke,
    )

    doc = SimpleDocTemplate(
        str(PDF_OUT), pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title="An Unofficial Catalogue of Visual Double Open Clusters",
        author="Alessandro Spina",
    )

    story = []
    story.append(Paragraph("An Unofficial Catalogue of Visual Double Open Clusters", title_style))
    story.append(Paragraph("Alessandro Spina &middot; deepskynotes.com &middot; First posted August 2026", subtitle_style))
    story.append(Paragraph(
        "This catalogue is a running list of the visual double open clusters I come across "
        "at the eyepiece. My working definition: any pair of open clusters that shares the "
        "same field of view in my 12-inch SkyWatcher Dob with a 35&nbsp;mm Panoptic eyepiece &mdash; "
        "a true field of about 1.6&deg;. Southern-hemisphere focused; a work in progress.",
        intro_style,
    ))
    story.append(Spacer(1, 6))

    # Table header + rows.
    def fmt_pair(p):
        return Paragraph(f"{p['pair'][0]}<br/>{p['pair'][1]}", cell_bold)

    def fmt_ra(p):
        return Paragraph(f"{p['ra_hms_a']}<br/>{p['ra_hms_b']}", cell_style)

    def fmt_dec(p):
        return Paragraph(f"{p['dec_dms_a']}<br/>{p['dec_dms_b']}", cell_style)

    def fmt_mag(p):
        b = "-" if p["mag_b"] is None else str(p["mag_b"])
        return Paragraph(f"{p['mag_a']}<br/>{b}", cell_style)

    def fmt_note(p):
        return Paragraph(p["note"], cell_style)

    header = [
        Paragraph("Pair", header_style),
        Paragraph("Constellation", header_style),
        Paragraph("RA (J2000)", header_style),
        Paragraph("Dec (J2000)", header_style),
        Paragraph("Mag", header_style),
        Paragraph("Observing notes", header_style),
    ]
    rows = [header]
    for p in PAIRS:
        rows.append([
            fmt_pair(p),
            Paragraph(p["const"], cell_style),
            fmt_ra(p),
            fmt_dec(p),
            fmt_mag(p),
            fmt_note(p),
        ])

    col_widths = [3.3 * cm, 2.4 * cm, 2.3 * cm, 2.3 * cm, 1.4 * cm, 15.5 * cm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.whitesmoke, colors.HexColor("#f9fafb")]),
    ]))
    story.append(tbl)

    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "Positions and magnitudes cross-checked between Stellarium and SIMBAD when each "
        "object entered the catalogue. Full observing notes, DSS plates and finder charts "
        "for each individual cluster are available at deepskynotes.com.",
        subtitle_style,
    ))

    doc.build(story)
    print(f"  {PDF_OUT.name}   {PDF_OUT.stat().st_size / 1024:6.1f} KB")


# ------------------------------------------------------------
# Excel: openpyxl workbook, header + data rows
# ------------------------------------------------------------
def build_xlsx() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Visual Doubles"

    headers = [
        "#", "Cluster A", "Cluster B", "Constellation",
        "RA A (J2000)", "Dec A (J2000)", "Mag A",
        "RA B (J2000)", "Dec B (J2000)", "Mag B",
        "Observing notes",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="374151")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_i, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for i, p in enumerate(PAIRS, start=1):
        row = [
            i,
            p["pair"][0], p["pair"][1], p["const"],
            p["ra_hms_a"], p["dec_dms_a"], p["mag_a"],
            p["ra_hms_b"], p["dec_dms_b"], p["mag_b"] if p["mag_b"] is not None else "-",
            p["note"],
        ]
        ws.append(row)
        for col_i in range(1, len(row) + 1):
            cell = ws.cell(row=i + 1, column=col_i)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_i == len(row)))
            cell.border = border

    widths = [4, 12, 12, 14, 14, 14, 8, 14, 14, 8, 90]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    for row_i in range(2, 2 + len(PAIRS)):
        ws.row_dimensions[row_i].height = 60

    wb.save(XLSX_OUT)
    print(f"  {XLSX_OUT.name}  {XLSX_OUT.stat().st_size / 1024:6.1f} KB")


if __name__ == "__main__":
    print("Building downloads:")
    build_pdf()
    build_xlsx()
    print(f"\nWrote to {OUT_DIR.relative_to(SITE_ROOT)}")
